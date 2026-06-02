import csv
import io
import math

import frappe
from frappe import _
from frappe.utils import cint

from scout.api.common import get_tpo_session_user
from scout.api.tpo.helpers import (
    MAX_CSV_EXPORT_ROWS,
    MAX_DIRECTORY_TOTAL,
    count_invites_without_matching_profile,
    count_student_profiles_by_filters,
    invite_rows_page_for_tpo,
    norm,
    student_rows_by_filters,
)

DEFAULT_STUDENT_PAGE_SIZE = 10
MAX_STUDENT_PAGE_SIZE = 50


def _serialize_profile_row(row: dict) -> dict:
    return {
        "studentId": row.get("student_user") or "",
        "fullName": row.get("full_name") or "",
        "email": row.get("email") or row.get("student_user") or "",
        "phone": row.get("phone") or "",
        "college": row.get("college") or "",
        "areaOfStudy": row.get("area_of_study") or "",
        "batch": row.get("academic_year") or "",
        "branch": row.get("department_stream") or "",
        "state": row.get("state") or "",
        "country": row.get("country") or "",
        "courseClassGrade": row.get("course_class_grade") or "",
        "resumeFile": row.get("resume_file") or "",
        "inviteStatus": "",
        "isPendingInvite": False,
    }


@frappe.whitelist(methods=["GET"])
def download_students_by_parameters():
    _user_id, err = get_tpo_session_user()
    if err:
        return err

    branch = norm(frappe.form_dict.get("branch"))
    batch = norm(frappe.form_dict.get("batch"))
    state = norm(frappe.form_dict.get("state"))
    country = norm(frappe.form_dict.get("country"))
    area_of_study = norm(frappe.form_dict.get("areaOfStudy"))

    rows = student_rows_by_filters(
        branch=branch,
        batch=batch,
        state=state,
        country=country,
        area_of_study=area_of_study,
        limit_page_length=MAX_CSV_EXPORT_ROWS,
    )

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["Student ID", "Name", "Email", "Phone", "College", "Area Of Study", "Batch", "Branch", "State", "Country", "Course/Class/Grade", "Resume"])
    for row in rows:
        writer.writerow(
            [
                row.get("student_user") or "",
                row.get("full_name") or "",
                row.get("email") or "",
                row.get("phone") or "",
                row.get("college") or "",
                row.get("area_of_study") or "",
                row.get("academic_year") or "",
                row.get("department_stream") or "",
                row.get("state") or "",
                row.get("country") or "",
                row.get("course_class_grade") or "",
                row.get("resume_file") or "",
            ]
        )

    frappe.local.response.filename = "students_filtered.csv"
    frappe.local.response.filecontent = stream.getvalue()
    frappe.local.response.type = "download"


def _pagination_meta(page: int, page_size: int, total: int, *, truncated: bool = False) -> dict:
    total_pages = max(1, math.ceil(total / page_size)) if total else 1
    page = min(page, total_pages)
    return {
        "page": page,
        "pageSize": page_size,
        "total": total,
        "totalPages": total_pages,
        "hasNext": page < total_pages,
        "hasPrev": page > 1,
        "truncated": truncated,
    }


@frappe.whitelist(methods=["GET"])
def count_students_directory():
    """Lightweight total for dashboard home (profiles + pending invites without profile)."""
    user_id, err = get_tpo_session_user()
    if err:
        return err

    branch = norm(frappe.form_dict.get("branch"))
    batch = norm(frappe.form_dict.get("batch"))
    state = norm(frappe.form_dict.get("state"))
    country = norm(frappe.form_dict.get("country"))
    area_of_study = norm(frappe.form_dict.get("areaOfStudy"))

    profile_total = count_student_profiles_by_filters(
        branch=branch,
        batch=batch,
        state=state,
        country=country,
        area_of_study=area_of_study,
    )
    invite_total = count_invites_without_matching_profile(
        user_id,
        branch=branch,
        batch=batch,
        state=state,
        country=country,
        area_of_study=area_of_study,
    )
    total = profile_total + invite_total
    return {
        "ok": True,
        "data": {
            "total": total,
            "profileCount": profile_total,
            "inviteCount": invite_total,
            "truncated": total >= MAX_DIRECTORY_TOTAL,
        },
    }


@frappe.whitelist(methods=["GET"])
def list_students_by_parameters():
    _user_id, err = get_tpo_session_user()
    if err:
        return err

    branch = norm(frappe.form_dict.get("branch"))
    batch = norm(frappe.form_dict.get("batch"))
    state = norm(frappe.form_dict.get("state"))
    country = norm(frappe.form_dict.get("country"))
    area_of_study = norm(frappe.form_dict.get("areaOfStudy"))

    page = max(1, cint(frappe.form_dict.get("page") or 1))
    page_size = cint(frappe.form_dict.get("pageSize") or DEFAULT_STUDENT_PAGE_SIZE)
    page_size = min(MAX_STUDENT_PAGE_SIZE, max(1, page_size))
    offset = (page - 1) * page_size

    profile_total = count_student_profiles_by_filters(
        branch=branch,
        batch=batch,
        state=state,
        country=country,
        area_of_study=area_of_study,
    )
    invite_total = count_invites_without_matching_profile(
        _user_id,
        branch=branch,
        batch=batch,
        state=state,
        country=country,
        area_of_study=area_of_study,
    )
    total = profile_total + invite_total
    truncated = total >= MAX_DIRECTORY_TOTAL

    students: list[dict] = []
    if offset < profile_total:
        profile_limit = min(page_size, profile_total - offset)
        rows = student_rows_by_filters(
            branch=branch,
            batch=batch,
            state=state,
            country=country,
            area_of_study=area_of_study,
            limit_start=offset,
            limit_page_length=profile_limit,
        )
        students.extend(_serialize_profile_row(row) for row in rows)

    remaining = page_size - len(students)
    if remaining > 0:
        invite_offset = max(0, offset - profile_total)
        students.extend(
            invite_rows_page_for_tpo(
                _user_id,
                branch=branch,
                batch=batch,
                state=state,
                country=country,
                area_of_study=area_of_study,
                limit_start=invite_offset,
                limit_page_length=remaining,
            )
        )

    pagination = _pagination_meta(page, page_size, total, truncated=truncated)
    return {"ok": True, "data": {"students": students, "pagination": pagination}}


def _normalize_header(value):
    return (value or "").strip().lower().replace("-", "_").replace(" ", "_")


EMAIL_HEADER_ALIASES = frozenset({"email", "e_mail", "student_email", "mail", "email_id"})


def _normalized_headers_from_first_row(rows):
    if not rows:
        return frozenset()
    return frozenset(_normalize_header(k) for k in rows[0].keys() if k)


def _sheet_has_email_column(rows):
    return bool(_normalized_headers_from_first_row(rows) & EMAIL_HEADER_ALIASES)


def _count_rows_with_email(rows):
    n = 0
    for raw_row in rows:
        mapped = _map_row_fields(raw_row)
        if norm(mapped.get("email")):
            n += 1
    return n


def _map_row_fields(raw_row):
    alias_map = {
        "email": {"email", "e_mail", "e-mail", "student_email", "mail", "email_id"},
        "full_name": {"full_name", "name", "student_name"},
        "batch": {"batch", "course_class_grade", "class_grade"},
        "year": {"year", "academic_year", "passout_year", "graduation_year"},
        "department": {"department", "department_stream", "branch", "stream"},
        "phone": {"phone", "mobile", "mobile_no", "contact"},
        "college": {"college", "college_name", "institute", "institute_name"},
        "area_of_study": {"area_of_study", "program", "course", "specialization"},
        "state": {"state"},
        "country": {"country"},
    }
    normalized = {_normalize_header(k): (v or "").strip() if isinstance(v, str) else (str(v).strip() if v is not None else "") for k, v in raw_row.items()}
    out = {}
    for key, aliases in alias_map.items():
        out[key] = ""
        for alias in aliases:
            if normalized.get(alias):
                out[key] = normalized.get(alias)
                break
    return out


def _parse_csv_or_excel(file_name, content):
    lower = (file_name or "").lower()
    if lower.endswith(".csv"):
        if isinstance(content, bytes):
            text = content.decode("utf-8-sig", errors="ignore")
        else:
            text = content.lstrip("﻿")  # strip BOM if str
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader if row]

    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        if isinstance(content, str):
            content = content.encode("utf-8")
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        parsed = []
        for row in rows[1:]:
            row_map = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                row_map[header] = row[i] if i < len(row) else ""
            if any((str(v).strip() for v in row_map.values() if v is not None)):
                parsed.append(row_map)
        return parsed

    frappe.throw(_("Only CSV and XLSX files are supported."))


@frappe.whitelist(methods=["POST"])
def bulk_upsert_students():
    from scout.api.tpo.bulk_upload import bulk_upsert_students_impl

    return bulk_upsert_students_impl()
